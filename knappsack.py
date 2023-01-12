from ortools.sat.python import cp_model
import numpy as np
import openpyxl as xl
from openpyxl.styles import PatternFill
import pprint as pp
filename = "TR1.csv"
resultsFile = filename[:-4] + "-Results.xlsx"
assignedColor = PatternFill(patternType="solid", start_color="81ef83")
costs = np.loadtxt(filename, dtype=int, delimiter=',')
wb = xl.Workbook()
ws = wb.active

for row in costs:
    ws.append(list(row))

# costs = 1/costs
costs = 10-costs
costs[costs==10] = 0
# with open(filename) as csvfile:
#     values = csv.reader(csvfile, delimiter=",")
#     for row in values:
#         costs.append(row)

num_students_per_lab = 2
pp.pp(costs)
print(type(costs[0][0]))
model = cp_model.CpModel()
num_students=len(costs)
num_labs = len(costs[0])

data = {}
data['weights'] = [1 for i in range(num_students) ]
data['costs'] = costs

assert len(data['costs']) == len(data['weights'])

data['num_students'] = num_students
data['all_students'] = range(data['num_students'])

data['lab_capacities'] = [num_students_per_lab for i in range(num_labs)]
data['num_labs'] = num_labs
data['all_labs'] = range(data['num_labs'])

# x[s, l] = 1 if student s is packed in lab l.
x = {}
for s in data['all_students']:
    for l in data['all_labs']:
        x[s, l] = model.NewBoolVar(f'x_{s}_{l}')

# Each student is assigned to at most one lab.
for s in data['all_students']:
    model.Add(sum(x[s, l] for l in data['all_labs']) <= 1)

# The amount packed in each lab cannot exceed its capacity.
for l in data['all_labs']:
    model.Add(
        sum(x[s, l] * data['weights'][s]
            for s in data['all_students']) <= data['lab_capacities'][l])

objective = []
for s in data['all_students']:
    for l in data['all_labs']:
        objective.append(
            cp_model.LinearExpr.Term(x[s,l], data['costs'][s,l])
        )
model.Maximize(cp_model.LinearExpr.Sum(objective))

solver = cp_model.CpSolver()
status = solver.Solve(model)

print(f'STATUS = {status}')

if status == cp_model.OPTIMAL:
    print(f'Total packed value: {solver.ObjectiveValue()}')
    total_weight = 0
    for l in data['all_labs']:
        print(f'Lab {l+1}')
        lab_weight = 0
        lab_value = 0
        for s in data['all_students']:
            if solver.Value(x[s, l]) > 0:
                print(
                    f"student {s} weight: {data['weights'][s]} value: {data['costs'][s][l]}"
                )
                lab_weight += data['weights'][s]
                lab_value += data['costs'][s][l]
                ws.cell(s+1,l+1).fill = assignedColor

        print(f'Packed lab weight: {lab_weight}')
        print(f'Packed lab value: {lab_value}\n')
        total_weight += lab_weight
    print(f'Total packed weight: {total_weight}')
else:
    print('The problem does not have an optimal solution.')

wb.save(resultsFile)